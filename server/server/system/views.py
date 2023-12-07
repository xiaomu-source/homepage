import os
from uuid import UUID

import inflection
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import HttpResponse, StreamingHttpResponse
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from rest_framework import serializers
from rest_framework.decorators import action

from system.models import Account, Role, Permission, UserOptLog, Resource
from system.serializers import AccountSerializer, RoleSerializer, PermissionSerializer, UserOptLogSerializer, \
    ResourceSerializer
from utils.authentications import CSRFExemptViewSet, AccountViewSet
from utils.decorator import action_record, schema_required, check_api_permission
from utils.files import get_media_delete_url, get_media_url
from utils.response import APIResponse, BadResponse
from utils.user import check_password, get_public_ip, make_password, get_email_avatar, parse_ip, get_browser
from utils.utils import get_jwt_token, dict_remove_empty_key, filter_queryset_by_partial_match, get_list_view_data
from utils.validations import StrField, CommaListField, EmailField, BoolField


class AuthViewSet(CSRFExemptViewSet):

    @schema_required(
        username=StrField(required=True, help_text="用户名"),
        password=StrField(required=True, help_text="密码"),
        code=serializers.IntegerField(required=True, help_text="验证码"),
    )
    @action(methods=["post"], detail=False, url_path="login")
    def login(self, request):
        # TODO 验证码校验
        username = self.validated_data.get("username")
        password = self.validated_data.get("password")
        account = Account.objects.filter(username=username)
        if not account:
            return BadResponse(message="用户名不存在.")
        account = account.first()
        if not check_password(password, account.password):
            return BadResponse(message="密码错误.")
        if not account.status:
            return BadResponse(message="当前账户已被禁用，请联系管理员.")
        user_data = {
            "_id": account.id,
            "username": account.username,
            "nickname": account.nickname,
            "roleId": account.role_id,
            "status": account.status,
        }
        user_data["token"] = get_jwt_token(user_data)
        return APIResponse(user_data, message="登录成功")

    @schema_required(
        username=StrField(required=True, help_text="用户名"),
        nickname=StrField(required=True, help_text="用户名"),
        password=StrField(required=True, help_text="密码"),
        email=EmailField(required=True, help_text="邮箱")
    )
    @action(methods=["post"], detail=False, url_path="register")
    def register(self, request):
        username = self.validated_data.get("username")
        nickname = self.validated_data.get("nickname")
        password = self.validated_data.get("password")
        email = self.validated_data.get("email")
        account = Account.objects.filter(username=self.validated_data.get("username"))
        if account:
            return BadResponse(message=f"用户名{username}已经注册")

        account_data = {
            "type": "admin",
            "avatar": get_email_avatar(email),
            "username": username,
            "nickname": nickname,
            "password": make_password(password),
            "userIp": (user_ip := get_public_ip(request)),
            "email": email,
            "address": parse_ip(user_ip),
            "platform": get_browser(request),
            "role": Role.objects.get(uid=UUID('a56b38e7-b16e-4236-9f03-1dc49ae49850')),
        }
        for key, value in account_data.copy().items():
            if not value:
                account_data.pop(key)
        # TODO 发送邮件
        Account(**account_data).save()
        return APIResponse(message="注册成功")

    @action(methods=["get"], detail=False, url_path="captcha")
    def captcha(self, request):
        # TODO 验证码逻辑
        data = '<svg xmlns="http://www.w3.org/2000/svg" width="90" height="38" viewBox="0,0,90,38"><path fill="#9651db" d="M16.44 34.01L16.52 34.08L16.52 34.08Q12.62 34.41 11.65 32.13L11.55 32.03L11.48 31.96Q12.41 31.21 13.95 29.46L13.88 29.39L13.90 29.41Q14.31 31.32 17.27 31.43L17.36 31.52L17.20 31.36Q20.63 31.54 22.24 30.23L22.13 30.13L22.13 30.13Q23.60 28.64 23.53 25.99L23.50 25.96L23.58 26.04Q23.32 20.96 18.61 21.18L18.58 21.15L18.69 21.26Q15.86 21.35 14.14 22.58L14.27 22.71L13.94 22.53L13.64 22.31L13.76 22.42Q14.09 19.31 13.98 16.62L13.85 16.50L13.83 16.48Q13.77 13.95 13.43 10.70L13.39 10.65L13.35 10.61Q17.07 11.68 20.81 11.57L20.77 11.53L20.87 11.62Q24.48 11.42 28.00 10.19L27.96 10.15L27.42 11.67L27.39 11.64Q27.07 12.36 26.88 13.19L26.99 13.30L26.96 13.26Q23.90 14.28 20.09 14.28L20.14 14.34L20.11 14.31Q18.63 14.24 17.17 14.10L17.16 14.08L17.13 14.06Q17.06 14.74 16.73 19.11L16.90 19.28L16.77 19.15Q17.65 18.76 19.67 18.61L19.72 18.67L19.60 18.54Q23.23 18.74 24.65 20.38L24.77 20.50L24.69 20.42Q26.12 22.00 26.38 25.78L26.33 25.72L26.42 25.81Q26.68 29.92 25.48 31.90L25.47 31.89L25.31 31.74Q23.39 33.67 20.14 33.89L20.10 33.85L20.12 33.87Q19.20 33.99 16.58 34.14ZM22.56 36.42L22.39 36.25L22.50 36.36Q25.92 36.49 27.79 35.41L27.79 35.41L27.80 35.42Q28.76 33.91 28.76 31.85L28.67 31.76L28.64 31.73Q28.71 28.77 27.77 24.73L27.66 24.62L27.68 24.65Q27.34 22.95 26.14 21.57L26.26 21.69L26.18 21.65L25.92 21.31L25.57 21.11L25.75 21.30Q25.38 20.51 24.75 19.84L24.81 19.90L24.71 19.84L24.74 19.87Q23.17 18.15 19.55 18.15L19.64 18.24L19.43 18.30L19.42 18.29Q19.37 17.60 19.56 16.52L19.54 16.50L19.70 16.66Q24.51 16.69 28.47 15.08L28.41 15.01L28.40 15.00Q28.90 13.71 29.83 10.94L29.76 10.87L27.71 11.81L27.78 11.88Q28.20 10.81 28.69 9.68L28.68 9.68L28.65 9.64Q24.90 11.10 20.87 11.17L20.90 11.21L20.82 11.12Q16.80 11.26 12.92 10.18L12.94 10.20L12.93 10.19Q13.62 14.55 13.62 18.89L13.59 18.85L13.50 18.76Q13.52 20.66 13.41 22.53L13.34 22.46L13.43 22.54Q13.67 22.71 14.23 23.01L14.22 23.00L14.16 22.94Q14.56 22.77 15.19 22.44L15.15 22.40L15.24 22.48Q15.26 23.29 15.04 24.79L14.92 24.67L15.07 24.82Q15.31 24.84 15.72 25.06L15.77 25.10L15.85 25.19Q18.61 23.49 20.44 23.49L20.44 23.49L20.35 23.40Q21.63 23.38 22.79 24.01L22.83 24.06L22.78 24.00Q23.10 24.99 23.13 25.93L23.25 26.05L23.25 26.05Q23.37 28.63 22.36 29.72L22.34 29.70L22.23 29.59Q20.87 30.77 18.66 30.99L18.63 30.96L18.79 31.12Q17.79 31.13 17.12 31.06L17.09 31.03L17.12 31.06Q16.31 30.96 15.56 30.62L15.47 30.53L15.49 30.36L15.32 30.41L15.42 30.52Q14.58 29.98 14.06 28.67L14.02 28.63L14.01 28.62Q12.80 30.10 11.15 32.19L11.08 32.12L11.13 32.17Q11.38 32.76 12.20 33.69L12.16 33.65L12.30 33.79Q13.25 35.49 16.02 35.97L16.12 36.08L16.02 35.98Q17.14 36.17 22.42 36.28Z"></path><path fill="#e2c15e" d="M66.82 21.05L66.85 21.08L66.87 21.10Q67.67 21.12 69.05 20.97L69.12 21.03L69.02 20.93Q69.04 21.63 69.04 22.27L68.97 22.19L69.06 23.52L69.11 23.57Q68.19 23.47 67.37 23.55L67.32 23.50L67.41 23.59Q66.49 23.53 65.67 23.49L65.83 23.65L65.69 23.52Q62.85 29.53 59.48 34.06L59.63 34.20L59.59 34.16Q57.13 34.81 55.86 35.41L55.86 35.41L55.89 35.43Q59.90 29.80 62.89 23.63L62.81 23.55L60.27 23.59L60.30 23.62Q60.28 22.29 60.17 20.98L60.26 21.07L60.08 20.90Q62.00 21.10 64.02 21.10L64.09 21.17L65.95 17.53L65.98 17.56Q66.94 15.67 68.13 14.10L68.16 14.13L68.12 14.09Q66.64 14.29 65.06 14.29L64.91 14.13L64.93 14.15Q59.02 14.23 55.32 12.06L55.25 11.99L54.62 10.31L54.66 10.35Q54.28 9.49 53.91 8.63L53.87 8.59L53.90 8.62Q58.30 11.33 63.90 11.56L63.73 11.39L63.77 11.43Q68.97 11.80 74.02 9.86L73.98 9.82L73.86 9.69Q73.72 10.15 73.23 11.01L73.29 11.07L73.23 11.01Q69.67 15.68 66.87 21.10ZM74.83 12.38L74.79 12.35L75.88 10.48L75.72 10.32Q74.92 10.95 73.28 11.66L73.24 11.62L73.39 11.29L73.48 11.37Q73.54 11.17 73.65 11.06L73.67 11.08L73.62 11.03Q73.97 10.33 74.68 9.03L74.80 9.14L74.68 9.02Q69.56 11.38 63.91 11.16L63.76 11.00L63.76 11.01Q57.86 10.75 53.19 7.83L53.35 7.99L53.35 7.99Q54.28 9.74 55.10 12.36L55.01 12.27L55.05 12.31Q56.21 13.02 57.00 13.32L57.01 13.34L56.90 13.23Q57.08 13.60 57.53 15.43L57.61 15.51L57.50 15.39Q60.67 16.70 66.10 16.55L65.98 16.44L66.15 16.61Q65.77 17.09 63.83 20.79L63.80 20.76L63.84 20.80Q61.76 20.74 59.81 20.55L59.96 20.70L59.77 20.51Q60.00 21.48 60.00 22.34L60.00 22.35L59.88 23.91L61.64 24.02L61.68 25.44L61.66 25.43Q57.69 32.56 54.96 36.08L55.09 36.20L55.03 36.15Q56.61 35.41 58.22 34.92L58.17 34.87L58.13 34.84Q57.43 35.70 56.08 37.39L56.08 37.39L56.17 37.47Q59.32 36.44 61.72 36.22L61.57 36.07L61.74 36.24Q64.53 32.45 67.59 25.79L67.70 25.90L71.06 26.08L70.89 25.91Q70.91 25.11 70.91 24.21L70.85 24.15L70.90 22.44L70.93 22.47Q70.59 22.40 70.05 22.44L70.15 22.53L70.03 22.41Q69.48 22.44 69.22 22.44L69.34 22.56L69.34 22.57Q69.30 22.41 69.34 22.26L69.36 22.28L69.27 21.94L69.39 22.05Q71.80 16.95 74.83 12.38Z"></path><path d="M2 15 C43 2,65 36,73 25" stroke="#d7d750" fill="none"></path><path fill="#6c46db" d="M42.58 33.23L42.53 33.17L42.59 33.24Q41.95 33.24 41.28 33.27L41.16 33.15L41.23 33.22Q40.47 33.25 39.80 33.25L39.77 33.23L39.79 33.24Q40.35 29.99 40.35 26.62L40.18 26.46L40.27 26.54Q38.47 26.58 37.58 26.58L37.66 26.66L37.67 26.67Q36.77 26.59 34.97 26.52L34.89 26.44L34.96 26.51Q34.91 26.16 34.73 23.73L34.80 23.80L34.88 23.88Q37.31 24.33 40.22 24.33L40.19 24.29L40.22 24.32Q40.03 20.06 39.62 17.52L39.66 17.55L39.66 17.56Q40.32 17.54 41.10 17.54L41.09 17.53L42.68 17.51L42.65 17.49Q42.55 21.79 42.55 24.33L42.51 24.29L42.55 24.33Q44.67 24.33 47.82 23.99L47.74 23.92L47.75 23.93Q47.62 25.21 47.62 26.41L47.73 26.52L47.62 26.41Q47.42 26.48 46.63 26.51L46.72 26.60L46.54 26.42Q45.59 26.44 44.99 26.48L45.11 26.60L44.99 26.47Q45.11 26.60 42.49 26.60L42.41 26.52L42.45 29.92L42.51 29.98Q42.39 31.46 42.50 33.15ZM48.21 23.49L48.27 23.55L48.20 23.48Q46.24 23.69 44.41 23.77L44.58 23.94L44.52 23.88Q44.78 20.81 45.19 18.98L45.22 19.01L45.14 18.92Q44.52 19.06 43.14 19.21L43.18 19.24L43.22 17.08L43.33 17.19Q40.72 17.16 39.08 17.05L39.07 17.04L39.12 17.09Q39.64 20.11 39.82 23.93L39.85 23.95L39.74 23.84Q38.13 23.92 34.51 23.32L34.38 23.20L34.48 23.30Q34.67 24.39 34.67 26.97L34.69 26.99L36.09 26.88L36.17 26.97Q36.12 27.63 35.97 29.01L36.00 29.04L39.97 28.86L39.81 28.70Q39.66 31.88 39.36 33.67L39.30 33.61L39.32 33.63Q40.08 33.65 41.46 33.53L41.46 33.53L41.50 33.57Q41.55 34.30 41.52 35.60L41.48 35.57L41.57 35.65Q41.94 35.51 45.49 35.66L45.60 35.76L45.53 35.69Q44.57 32.60 44.35 28.71L44.44 28.81L44.50 28.87Q47.61 28.79 49.66 29.17L49.76 29.26L49.65 29.16Q49.51 28.27 49.51 27.33L49.48 27.30L49.54 25.49L49.48 25.43Q49.28 25.46 48.76 25.50L48.80 25.54L48.00 25.52L48.06 25.58Q48.18 24.92 48.29 23.57Z"></path></svg>'

        return APIResponse(data)


class PermissionViewSet(AccountViewSet):
    @schema_required(with_list=True)
    @check_api_permission('sys:permissions:list')
    @action_record(module="权限/列表")
    @action(methods=["post"], detail=False, url_path="get-list")
    def get_list(self, request):
        try:
            model_objs = filter_queryset_by_partial_match(Permission, self.validated_data.get("params"))
            if order_by_filed := self.validated_data.get("order_by_field"):
                model_objs = model_objs.order_by(order_by_filed)
            permission_tree = Permission.to_tree(model_objs)
            return APIResponse({"result": PermissionSerializer(permission_tree, many=True).data})
        except Exception as e:
            raise e

    @action(methods=["post"], detail=False, url_path="tree")
    def tree(self, request):
        try:
            permissions = Permission.objects.all().order_by("sort_order")
            permission_tree = Permission.to_tree(permissions)
            return APIResponse({"result": PermissionSerializer(permission_tree, many=True).data})
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        under_sore=True,
        name=StrField(required=True, help_text="权限名称"),
        key=StrField(required=True, help_text="权限标识"),
        parent_key=StrField(required=False, help_text="父级标识"),
        auth=BoolField(required=False, help_text="是否是权限按钮", default=False),
        autoSon=BoolField(required=False, help_text="是否生成子级", default=False),
        sortOrder=serializers.IntegerField(required=False, help_text="排序", default=0)
    )
    @check_api_permission('sys:permissions:create')
    @action_record(module="系统管理/权限管理/新增")
    @action(methods=["post"], detail=False, url_path="create-permission")
    def create_permission(self, request):
        name = self.validated_data.get('name')
        key = self.validated_data.get('key')
        parent_key = self.validated_data.get('parent_key')
        auth = self.validated_data.get('auth') or False
        auto_son = self.validated_data.get('auto_son') or False
        sort_order = self.validated_data.get('sort_order') or 0

        new_permissions = [
            {'name': name, 'key': key, 'parent_key': parent_key, 'auth': auth, "sort_order": sort_order,}
        ]
        if key.count(":") == 1 and auto_son:
            new_permissions.extend([
                {'name': '查询', 'key': f'{key}:list', 'parent_key': key, 'auth': auth},
                {'name': '增加', 'key': f'{key}:create', 'parent_key': key, 'auth': auth},
                {'name': '删除', 'key': f'{key}:delete', 'parent_key': key, 'auth': auth},
                {'name': '更新', 'key': f'{key}:update', 'parent_key': key, 'auth': auth},
            ])
        Permission.objects.bulk_create([Permission(**permission) for permission in new_permissions])
        data = {'name': name, 'key': key, 'parent_key': parent_key, 'auth': auth, "autoSon": auto_son,"sortOrder": sort_order,}
        return APIResponse(data, message="添加权限成功")

    @schema_required(
        with_id=True,
        under_sore=True,
        name=StrField(required=True, help_text="权限名称"),
        key=StrField(required=True, help_text="权限标识"),
        parent_key=StrField(required=False, help_text="父级标识"),
        auth=BoolField(required=False, help_text="是否是权限按钮", default=False),
        autoSon=BoolField(required=False, help_text="是否生成子级", default=False),
        sortOrder=serializers.IntegerField(required=False, help_text="排序", default=0)
    )
    @check_api_permission('sys:permissions:update')
    @action_record(module="系统管理/权限管理/更新")
    @action(methods=["post"], detail=False, url_path="update-permission")
    def update_permission(self, request):
        try:
            permission = Permission.objects.filter(id=self.validated_data.get("id"))
            if not permission:
                return BadResponse(message="该权限管理不存在")
            update_data = {_key: value for _key, value in self.validated_data if _key not in ["id", "auto_son"] and value}
            permission.update(**update_data)
            return APIResponse(message="权限管理更新成功")
        except Exception as e:
            return BadResponse(str(e), message="权限更新失败")

    @schema_required(
        with_id=True,
        key=StrField(required=True, help_text="权限标识"),
    )
    @check_api_permission('sys:permissions:delete')
    @action_record(module="系统管理/权限管理/删除")
    @action(methods=["post"], detail=False, url_path="delete-permission")
    def delete_permission(self, request):
        _id = self.validated_data.get("id")
        key = self.validated_data.get("key")

        if key.count(":") <= 1:
            Permission.objects.filter(key__istartswith=key).delete()
            return APIResponse(message="权限删除成功")
        Permission.objects.filter(id=_id).delete()
        return APIResponse(message="权限删除成功")

    @schema_required(with_id=True,)
    @check_api_permission('sys:permissions:stop')
    @action_record(module="系统管理/权限管理/停用权限")
    @action(methods=["post"], detail=False, url_path="stop")
    def stop(self, request):
        permission = Permission.objects.filter(id=self.validated_data.get("id"))
        if not permission:
            return BadResponse("该条数据不存在")
        data = dict(
            status = not permission.status,
            disabled = not permission.disabled,
        )
        permission.update(**data)
        return APIResponse(message="权限管理更新成功")


class ResourceViewSet(AccountViewSet):
    @schema_required(with_list=True)
    @check_api_permission('sys:resources:list')
    @action(methods=["post"], detail=False, url_path="get-list")
    def get_list(self, request):
        try:
            return APIResponse(get_list_view_data(Resource, ResourceSerializer, self.validated_data))
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        under_sore=True,
        srcName = StrField(required=True, help_text="资源名称"),
        deletePath = StrField(required=True, help_text="资源类型"),
        downloadPath = StrField(required=True, help_text="资源下载路径"),
        previewPath = StrField(required=True, help_text="资源预览路径"),
        srcType = StrField(required=True, help_text="资源类型"),
        srcSize = StrField(required=True, help_text="资源大小"),
        remark = StrField(required=False, help_text="备注"),
    )
    @check_api_permission('sys:resources:create')
    @action_record("资源管理/创建")
    @action(methods=["post"], detail=False, url_path="create-resource")
    def create_resource(self, request):
        media_url = get_media_url(self.validated_data.get("src_name"))
        media_delete_url = get_media_delete_url(self.validated_data.get("src_name"))
        self.validated_data["delete_path"] = media_delete_url
        self.validated_data["download_path"] = media_url
        self.validated_data["preview_path"] = media_url
        self.validated_data["user"] = request.user
        resource = Resource.objects.create(**self.validated_data)
        return APIResponse(data = ResourceSerializer(resource).data, message="添加资源管理成功.")

    @schema_required(
        with_id=True,
        srcName=StrField(required=True, help_text="资源名称"),
        deletePath=StrField(required=True, help_text="资源类型"),
        downloadPath=StrField(required=True, help_text="资源下载路径"),
        previewPath=StrField(required=True, help_text="资源预览路径"),
        srcType=StrField(required=True, help_text="资源类型"),
        srcSize=StrField(required=True, help_text="资源大小"),
        remark=StrField(required=False, help_text="备注"),
    )
    @check_api_permission('sys:resources:update')
    @action_record("资源管理/更新")
    @action(methods=["post"], detail=False, url_path="update-resource")
    def update_resource(self, request):
        data = {inflection.underscore(key): value for key, value in self.validated_data.items() if value or key not in ["id", "_id"]}
        media_url = get_media_url(self.validated_data.get("src_name"))
        media_delete_url = get_media_delete_url(self.validated_data.get("src_name"))
        data["delete_path"] = media_delete_url
        data["download_path"] = media_url
        data["preview_path"] = media_url
        data["user"] = request.user
        Resource.objects.filter(id=self.validated_data.get("id")).update(**data)
        return APIResponse(message="资源管理更新成功.")

    @schema_required(with_id=True,)
    @check_api_permission('sys:resources:delete')
    @action_record("资源管理/删除")
    @action(methods=["post"], detail=False, url_path="delete-resource")
    def delete_resource(self, request):
        try:
            resource = Resource.objects.get(id=self.validated_data.get("id"))
            file_path = os.path.join(settings.MEDIA_ROOT, resource.src_name)
            resource.delete()
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
            return APIResponse(message="删除资源管理成功")
        except Resource.DoesNotExist:
            return BadResponse(message="该资源不存在")
        except Exception as e:
            return BadResponse(str(e))


class RoleViewSet(AccountViewSet):
    @schema_required(with_list=True)
    @check_api_permission('sys:roles:list')
    @action(methods=["post"], detail=False, url_path="get-list")
    def get_list(self, request):
        try:
            return APIResponse(get_list_view_data(Role, RoleSerializer, self.validated_data))
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        under_sore=True,
        roleName=StrField(required=False, help_text="角色名称"),
        roleAuth=StrField(required=False, help_text="角色标识"),
        perms=CommaListField(required=False, help_text="权限类别"),
        remark=StrField(required=False, help_text="备注"),
        status=BoolField(required=False, help_text="状态: 0禁用 1正常"),
    )
    @check_api_permission('sys:roles:list')
    @action_record("系统管理/角色管理/新增")
    @action(methods=["post"], detail=False, url_path="create-role")
    def create_role(self, request):
        role_data = {
            "role_name": self.validated_data.get("role_name"),
            "role_auth": self.validated_data.get("role_auth"),
            "perms": self.validated_data.get("perms"),
            "remark": self.validated_data.get("remark"),
            "status": self.validated_data.get("status") or True,
        }
        Role(**dict_remove_empty_key(role_data)).save()
        return APIResponse(message="创建角色成功")

    @schema_required(
        with_id=True,
        under_sore=True,
        roleName=StrField(required=False, help_text="角色名称"),
        roleAuth=StrField(required=False, help_text="角色标识"),
        perms=CommaListField(required=False, help_text="权限类别"),
        remark=StrField(required=False, help_text="备注"),
        status=BoolField(required=False, help_text="状态: 0禁用 1正常"),
    )
    @check_api_permission('sys:roles:update')
    @action_record("系统管理/角色管理/更新")
    @action(methods=["post"], detail=False, url_path="update-role")
    def update_role(self, request):
        try:
            role = Role.objects.get(pk=self.validated_data.get("id"))
            if role.role_auth == "SUPER-ADMIN":
                return BadResponse(message="不能修改超级管理员")
            for key, value in request.data.items():
                if key not in ["id"] and value:
                    setattr(role, key, value)
            role.save()
            return APIResponse("角色更新成功.")
        except Role.DoesNotExist:
            return BadResponse("该角色不存在")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(with_id=True)
    @check_api_permission('sys:roles:update')
    @action_record("系统管理/角色管理/删除")
    @action(methods=["post"], detail=False, url_path="delete-role")
    def delete_role(self, request):
        try:
            role = Role.objects.get(pk=self.validated_data.get("id"))
            if role.role_auth == "SUPER-ADMIN":
                return BadResponse(message="不能删除超级管理员角色")
            if role.role_auth == "VISITOR-ADMIN":
                return BadResponse(message="不能删除访客管理员角色")
            role.delete()
            return APIResponse("删除角色成功.")
        except Role.DoesNotExist:
            return BadResponse("该角色不存在或已被删除")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(with_id=True)
    @action(methods=["post"], detail=False, url_path="findOne")
    def find_one(self, request):
        try:
            role = Role.objects.get(id=self.validated_data.get("id"))
            return APIResponse(RoleSerializer(role).data, message="查询成功.")
        except Role.DoesNotExist:
            return BadResponse(message="该角色不存在")
        except Exception as e:
            return BadResponse(str(e))


class UserViewSet(AccountViewSet):
    @schema_required(with_list=True)
    @check_api_permission("sys:users:list")
    @action(methods=["post"], detail=False, url_path="get-list")
    def get_list(self, request):
        try:
            return APIResponse(get_list_view_data(Account, AccountSerializer, self.validated_data))
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        under_sore=True,
        avatar = StrField(required=True, help_text="用户头像"),
        username=StrField(required=True, help_text="用户名"),
        nickname=StrField(required=True, help_text="用户名"),
        password=StrField(required=True, help_text="密码"),
        email=EmailField(required=False, help_text="邮箱"),
        remark=StrField(required=False, help_text="备注"),
        roleId=serializers.IntegerField(required=True, help_text="角色ID"),
    )
    @check_api_permission("sys:users:create")
    @action_record("系统管理/用户管理/新增")
    @action(methods=["post"], detail=False, url_path="create-user")
    def create_user(self, request):
        avatar = self.validated_data.get("avatar")
        username = self.validated_data.get("username")
        nickname = self.validated_data.get("nickname")
        password = self.validated_data.get("password")
        email = self.validated_data.get("email")
        remark = self.validated_data.get("remark")
        role_id = self.validated_data.get("role_id")
        if Account.objects.filter(username=self.validated_data.get("username")):
            return BadResponse(message=f"用户名{username}已经注册")

        account_data = {
            "type": "admin",
            "avatar": avatar or get_email_avatar(email),
            "username": username,
            "nickname": nickname,
            "password": make_password(password),
            "userIp": (user_ip := get_public_ip(request)),
            "email": email,
            "address": parse_ip(user_ip),
            "platform": get_browser(request),
            "remark": remark,
            "role": Role.objects.get(id=role_id),
        }
        for key, value in account_data.copy().items():
            if not value:
                account_data.pop(key)
        # TODO 发送邮件
        account = Account.objects.create(**account_data)
        return APIResponse(AccountSerializer(account).data, message="添加管理员成功")

    @schema_required(
        with_id=True,
        under_sore=True,
        avatar=StrField(required=True, help_text="用户头像"),
        username=StrField(required=True, help_text="用户名"),
        nickname=StrField(required=True, help_text="用户名"),
        password=StrField(required=True, help_text="密码"),
        email=EmailField(required=False, help_text="邮箱"),
        remark=StrField(required=False, help_text="备注"),
        roleId=serializers.IntegerField(required=True, help_text="角色ID"),
    )
    @check_api_permission("sys:users:update")
    @action_record("系统管理/用户管理/新增")
    @action(methods=["post"], detail=False, url_path="update-user")
    def update_user(self, request):
        try:
            account = Account.objects.get(pk=self.validated_data.get("id"))
            if role_id := self.validated_data.get("role_id"):
                self.validated_data["role"] = Role.objects.get(id=role_id)
                self.validated_data.pop("role_id")
            for key, value in self.validated_data.items():
                if key not in ["id"] and value:
                    setattr(account, key, value)
            account.save()
            return APIResponse(message="更新成功.")
        except Account.DoesNotExist:
            return BadResponse(message="该用戶不存在")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(with_id=True)
    @check_api_permission("sys:users:delete")
    @action_record("系统管理/用户管理/删除")
    @action(methods=["post"], detail=False, url_path="delete-user")
    def delete_user(self, request):
        try:
            account = Account.objects.filter(pk=self.validated_data.get("id"))
            if not account:
                return BadResponse(message="该管理员不存在或已被删除")
            account.delete()
            return APIResponse(message="删除管理员成功.")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        with_id=True,
        password=StrField(required=False,help_text="新密码")
    )
    @check_api_permission("sys:users:reset")
    @action_record("系统管理/用户管理/重置密码")
    @action(methods=["post"], detail=False, url_path="reset")
    def reset(self, request):
        try:
            _id = self.validated_data.get("id")
            account = Account.objects.get(pk=_id)
            password = self.validated_data.get("password")
            account.password = make_password(password)
            account.save()
            return APIResponse({"_id": _id , "password": password},message="重置成功")
        except Account.DoesNotExist:
            return BadResponse(message="该管理员不存在")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(with_id=True)
    @action(methods=["post"], detail=False, url_path="findOne")
    def find_one(self, request):
        try:
            account = Account.objects.get(id=self.validated_data.get("id"))
            return APIResponse(AccountSerializer(account).data, message="成功")
        except Account.DoesNotExist:
            return BadResponse(message="该管理员不存在")
        except Exception as e:
            return BadResponse(str(e))


class UserOptLogViewSet(AccountViewSet):
    @schema_required(with_list=True)
    @action(methods=["post"], detail=False, url_path="get-list")
    def get_list(self, request):
        try:
            return APIResponse(get_list_view_data(UserOptLog, UserOptLogSerializer, self.validated_data))
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        operator = StrField(required=True, help_text="操作人"),
        module = StrField(required=True, help_text="操作模块"),
        platform = StrField(required=True, help_text="操作平台"),
        operatorIP = StrField(required=True, help_text="设备IP"),
        address = StrField(required=True, help_text="设备位置"),
        content = StrField(required=False, help_text="操作内容"),
    )
    @check_api_permission("sys:users_opt_logs:create")
    @action_record("系统管理/操作日志/创建")
    @action(methods=["post"], detail=False, url_path="create-log")
    def create_log(self, request):
        log_data = dict(
            operator=self.validated_data.get("operator"),
            module=self.validated_data.get("module"),
            platform=self.validated_data.get("platform"),
            operator_ip=self.validated_data.get("operatorIP") or "0.0.0.0",
            address=self.validated_data.get("address") or "-",
            content=self.validated_data.get("content"),
        )
        user_opt_log = UserOptLog(**log_data)
        user_opt_log.save()
        return APIResponse(data=log_data, message="添加操作日志成功.")

    @schema_required(
        with_id=True,
        operator=StrField(required=True, help_text="操作人"),
        module=StrField(required=True, help_text="操作模块"),
        platform=StrField(required=True, help_text="操作平台"),
        operatorIP=StrField(required=True, help_text="设备IP"),
        address=StrField(required=True, help_text="设备位置"),
        content=StrField(required=False, help_text="操作内容"),
    )
    @check_api_permission("sys:users_opt_logs:update")
    @action_record("系统管理/操作日志/更新")
    @action(methods=["post"], detail=False, url_path="update-log")
    def update_log(self, request):
        target = UserOptLog.objects.get(id=self.validated_data.get("id"))
        for key, value in self.validated_data.items():
            if key not in ["id", "_id"]:
                setattr(target, key, value)
        target.save()
        return APIResponse(message="操作日志更新成功")

    @schema_required(with_id=True)
    @check_api_permission("sys:users_opt_logs:delete")
    @action_record("系统管理/操作日志/删除")
    @action(methods=["post"], detail=False, url_path="delete-log")
    def delete_log(self, request):
        try:
            UserOptLog.objects.get(id=self.validated_data.get("id")).delete()
            return APIResponse(message="删除操作日志成功")
        except UserOptLog.DoesNotExist:
            return BadResponse("该操作日志不存在或已被删除")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(
        ids = CommaListField(default=True, help_text="id 列表")
    )
    @check_api_permission("sys:users_opt_logs:deleteAll")
    @action_record("系统管理/操作日志/批量删除")
    @action(methods=["post"], detail=False, url_path="deleteAll")
    def delete_all(self, request):
        try:
            UserOptLog.objects.filter(id__in=self.validated_data.get("ids")).delete()
            return APIResponse(message="批量删除操作日志成功")
        except Exception as e:
            return BadResponse(str(e))

    @schema_required(with_list=True)
    @check_api_permission("sys:users_opt_logs:export")
    @action_record("系统管理/操作日志/导出")
    @action(methods=["post"], detail=False, url_path="export")
    def export(self, request):
        """TODO fix"""
        user_opt_logs = filter_queryset_by_partial_match(UserOptLog, self.validated_data.get("params")).order_by(self.validated_data.get("order_by_field"))
        wb = Workbook()
        ws = wb.active
        headers = ["operator", "operatorId", "module", "platform", "operatorIP", "content"]
        ws.append(headers)
        for log in user_opt_logs:
            row_data = [log.operator, log.operator_id, log.module, log.platform, log.operator_ip, log.content, ]
            ws.append(row_data)
        file_name = "操作日志.xlsx"
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        wb.save(file_path)
        wb.close()
        response = StreamingHttpResponse(open(file_path, "rb"))
        response["Content-Type"] = "application/octet-stream"
        response["Content-Disposition"] = f"attachment;filename={file_name}"
        return response

    @check_api_permission("sys:users_opt_logs:import")
    @action_record("系统管理/操作日志/导入")
    @action(methods=["post"], detail=False, url_path="import")
    def import_log(self, request):
        """TODO fix"""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return BadResponse("没有上传文件")
        file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.name)
        # 保存之后名字可能会改变，重新获取一下
        file_name = default_storage.save(file_path, ContentFile(uploaded_file.read()))

        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        workbook = load_workbook(file_path)

        sheet = workbook.active

        # Iterate through rows and columns to read data
        objs = []
        for row in range(1, sheet.max_row + 1):
            if not sheet.cell(row=row, column=1).value:
                continue
            log_data = dict(
                operator=sheet.cell(row=row, column=1).value,
                module=sheet.cell(row=row, column=2).value,
                platform=sheet.cell(row=row, column=3).value,
                operator_ip=sheet.cell(row=row, column=4).value or "0.0.0.0",
                address=sheet.cell(row=row, column=5).value or "-",
                content=sheet.cell(row=row, column=6).value or "",
            )
            objs.append(UserOptLog(**log_data))
        # Close the workbook when done
        workbook.close()
        default_storage.delete(file_path)
        UserOptLog.objects.bulk_create(objs, ignore_conflicts=True)
        return APIResponse(message="操作日志导入成功")

    @action(methods=["post"], detail=False, url_path="downloadTemplate")
    def download_template(self, request):
        template_name = "操作日志导入模板.xlsx"
        template_file_path = os.path.join(settings.STATIC_ROOT, template_name)
        response = StreamingHttpResponse(open(template_file_path, "rb"))
        response["Content-Type"] = "application/octet-stream"
        response["Content-Disposition"] = f"attachment;filename={template_name}"
        return response
